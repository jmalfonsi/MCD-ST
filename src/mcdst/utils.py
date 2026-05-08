from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import unicodedata
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

import yaml


class NoAliasSafeDumper(yaml.SafeDumper):
    def ignore_aliases(self, data):
        return True


CSV_ENCODINGS = ("utf-8-sig", "cp1252", "latin-1")
CSV_DELIMITERS = ",;\t|"
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
SOURCE_SHEET_SEPARATOR = "#"
SUPPORTED_EXPORT_SUFFIXES = {".csv", *EXCEL_SUFFIXES}


def utc_now() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")


def normalize(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", str(value))
    ascii_only = "".join(char for char in decomposed if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]+", "", ascii_only.lower())


def normalize_upper(value: str) -> str:
    if not value:
        return ""
    text = unicodedata.normalize("NFKD", value)
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"[^A-Za-z0-9]+", "_", text).strip("_")
    return text.upper()


def hash_id(prefix: str, value: str, salt: str = "mcdst-local") -> str:
    digest = hashlib.sha256(f"{salt}|{value}".encode("utf-8")).hexdigest()[:10]
    return f"{prefix}_{digest}"


def non_empty(values: list[str]) -> int:
    return sum(1 for value in values if value.strip())


def first_examples(values: list[str], limit: int = 3) -> list[str]:
    seen = []
    for value in values:
        if value and value not in seen:
            seen.append(value)
        if len(seen) == limit:
            break
    return seen


def is_year(value: str) -> bool:
    return bool(re.match(r"^(19|20)[0-9]{2}$", value))


def is_date(value: str) -> bool:
    if not value:
        return False
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            datetime.strptime(value, fmt)
            return True
        except ValueError:
            pass
    return False


def is_number(value: str) -> bool:
    try:
        float(value.replace(",", "."))
        return True
    except ValueError:
        return False


def list_tabular_sources(path: Path) -> list[dict[str, str | None]]:
    sources = []
    for file_path in sorted(path.iterdir()):
        if not file_path.is_file() or file_path.name.startswith("~$"):
            continue
        if file_path.suffix.lower() == ".csv":
            sources.append(
                {
                    "source_ref": file_path.name,
                    "file": file_path.name,
                    "sheet": None,
                    "format": "csv",
                }
            )
        elif file_path.suffix.lower() in EXCEL_SUFFIXES:
            for sheet_name in list_excel_sheets(file_path):
                sources.append(
                    {
                        "source_ref": format_source_ref(file_path.name, sheet_name),
                        "file": file_path.name,
                        "sheet": sheet_name,
                        "format": "excel",
                    }
                )
    return sources


def format_source_ref(filename: str, sheet_name: str | None = None) -> str:
    if not sheet_name:
        return filename
    return f"{filename}{SOURCE_SHEET_SEPARATOR}{sheet_name}"


def split_source_ref(source_ref: str) -> tuple[str, str | None]:
    if SOURCE_SHEET_SEPARATOR not in source_ref:
        return source_ref, None
    filename, sheet_name = source_ref.split(SOURCE_SHEET_SEPARATOR, 1)
    return filename, sheet_name


def read_source_rows(exports_dir: Path, source_ref: str) -> list[dict[str, str]]:
    filename, sheet_name = split_source_ref(source_ref)
    path = exports_dir / filename
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in EXCEL_SUFFIXES:
        if not sheet_name:
            raise ValueError(f"Excel source reference must include a sheet name: {source_ref}")
        return read_excel_sheet(path, sheet_name)
    raise ValueError(f"Unsupported export format for {source_ref}. Supported formats: csv, xlsx, xlsm.")


def read_csv(path: Path) -> list[dict[str, str]]:
    last_error = None
    for encoding in CSV_ENCODINGS:
        try:
            text = path.read_text(encoding=encoding)
            return read_csv_text(text)
        except UnicodeDecodeError as exc:
            last_error = exc
    supported = ", ".join(CSV_ENCODINGS)
    raise UnicodeError(f"Could not decode CSV {path} with supported encodings: {supported}") from last_error


def read_csv_text(text: str) -> list[dict[str, str]]:
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=CSV_DELIMITERS)
    except csv.Error:
        dialect = csv.excel
    return normalize_rows(csv.DictReader(io.StringIO(text), dialect=dialect))


def list_excel_sheets(path: Path) -> list[str]:
    workbook = load_excel_workbook(path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_excel_sheet(path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_excel_workbook(path, read_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_index = next(
        (index for index, row in enumerate(rows) if any(cell_to_text(cell) for cell in row)),
        None,
    )
    if header_index is None:
        return []

    headers = [cell_to_text(cell) for cell in rows[header_index]]
    indexed_headers = [(index, header) for index, header in enumerate(headers) if header]
    output = []
    for row in rows[header_index + 1 :]:
        if not any(cell_to_text(cell) for cell in row):
            continue
        output.append(
            {
                header: cell_to_text(row[index]) if index < len(row) else ""
                for index, header in indexed_headers
            }
        )
    return output


def load_excel_workbook(path: Path, *, read_only: bool):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Excel exports require the optional dependency openpyxl. "
            "Install the package with `python3 -m pip install -e .`."
        ) from exc
    return load_workbook(path, read_only=read_only, data_only=True)


def normalize_rows(rows: csv.DictReader) -> list[dict[str, str]]:
    return [
        {str(key): cell_to_text(value) for key, value in row.items() if key is not None}
        for row in rows
    ]


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == datetime.min.time() else value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_yaml(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        yaml.dump(payload, Dumper=NoAliasSafeDumper, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )


def read_yaml(path: Path) -> Any:
    return yaml.safe_load(path.read_text(encoding="utf-8"))
