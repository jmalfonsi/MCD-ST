from __future__ import annotations

import csv
import json
import threading
from contextlib import contextmanager
from pathlib import Path
from shutil import copyfile
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from mcdst.api import create_server
from mcdst.cli import main
from mcdst.cohort import evaluate_cohort_definition
from mcdst.engine import apply_mapping_file, apply_review_workdir, propose_mapping_workdir
from mcdst.learning import build_column_suggestions, build_learning_dataset
from mcdst.profiling import profile_exports
from mcdst.registry import load_registry
from mcdst.utils import write_yaml


FIXTURE_EXPORTS = Path(__file__).parent / "fixtures" / "mapping_poc_exports"
SEMICOLON_CSV_FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "semicolon_csv_export"
    / "salaries_suivis_small.csv"
)
COHORT_FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "cohorts"
    / "travailleurs_45_plus_manutention.yaml"
)
LONGITUDINAL_COHORT_FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "cohorts"
    / "manutention_avant_restriction.yaml"
)
FIXTURE_REVIEW_DECISIONS = Path(__file__).parent / "fixtures" / "mapping_poc_review_decisions.yaml"
CANONICAL_EXPORT_FILES = (
    "export_01_individus.csv",
    "export_02_structures.csv",
    "export_03_actes.csv",
    "export_04_risques.csv",
    "export_05_biometrie.csv",
    "export_06_pathologies_atmp.csv",
    "export_07_arrets.csv",
    "export_08_vaccinations.csv",
    "export_09_duerp.csv",
)
MAPPING_CONTRACT_KEYS = {
    "mapping_id",
    "source_system",
    "schema_version",
    "mapping_version",
    "review_status",
    "generated_at",
    "thresholds",
    "source_files",
    "join_candidates",
    "join_rules",
    "alternate_entity_sources",
    "entities",
    "review_queue",
    "value_mappings",
    "blocked_fields",
}
ENTITY_CONTRACT_KEYS = {"source_file", "review_status", "fields"}
FIELD_CONTRACT_KEYS = {
    "source",
    "confidence_score",
    "sensitivity",
    "transform",
    "review_status",
}


def test_mapping_poc_fixture_directory_contains_only_canonical_exports():
    assert tuple(sorted(path.name for path in FIXTURE_EXPORTS.glob("*.csv"))) == CANONICAL_EXPORT_FILES


def test_mapping_engine_roundtrip(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    output = tmp_path / "mcdst_tables"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )

    mapping = proposed["mapping"]
    assert len(mapping["entities"]) == 13
    assert len(mapping["blocked_fields"]) == 6
    assert len(mapping["join_candidates"]) == 31
    assert len(mapping["join_rules"]) == 31
    assert len(mapping["review_queue"]) == 2
    assert proposed["dry_run"]["quality"]["summary"]["generated_tables"]["travailleur"] == 4

    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_column_decisions(proposed["review_queue"]))

    validated = apply_review_workdir(workdir, decisions_path)
    assert len(validated["review_queue"]) == 0

    state = apply_mapping_file(workdir / "mapping_valide.yaml", exports, output)
    summary = state["quality"]["summary"]
    assert summary["generated_tables"] == {
        "travailleur": 4,
        "entreprise": 3,
        "etablissement": 3,
        "unite_travail": 4,
        "episode_poste": 4,
        "visite_sante_travail": 4,
        "conclusion_medicale": 4,
        "exposition_professionnelle": 4,
        "examen_complementaire": 4,
        "pathologie_atmp": 4,
        "arret_travail": 4,
        "vaccination": 4,
        "risque_unite_travail": 4,
    }
    assert summary["blocked_fields_count"] == 6
    assert summary["review_queue_count"] == 0
    assert summary["join_candidates_count"] == 31
    assert summary["join_rules_count"] == 31
    assert summary["value_mapping_groups_count"] == 11

    pending_values = [
        rule for rule in state["quality"]["rules"]
        if rule["rule"] == "review_queue:pending_value_mappings"
    ][0]
    assert pending_values["status"] == "failed"
    assert "4 nomenclature values" in pending_values["message"]


def test_mapping_engine_applies_value_review_decisions(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    output = tmp_path / "mcdst_tables"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )

    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_review_decisions(proposed["review_queue"]))

    validated = apply_review_workdir(workdir, decisions_path)
    assert len(validated["review_queue"]) == 0
    assert not [
        item
        for group in validated["value_mappings"]
        for item in group["mappings"]
        if item["status"] == "a_revoir"
    ]

    state = apply_mapping_file(workdir / "mapping_valide.yaml", exports, output)
    pending_values = [
        rule for rule in state["quality"]["rules"]
        if rule["rule"] == "review_queue:pending_value_mappings"
    ][0]
    assert pending_values["status"] == "passed"
    assert state["quality"]["summary"]["review_queue_count"] == 0


def test_cohort_engine_evaluates_yaml_definition(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    output = tmp_path / "mcdst_tables"
    report_path = tmp_path / "cohort_report.json"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )
    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_review_decisions(proposed["review_queue"]))
    apply_review_workdir(workdir, decisions_path)
    apply_mapping_file(workdir / "mapping_valide.yaml", exports, output)

    result = evaluate_cohort_definition(output, COHORT_FIXTURE, report_path)

    assert result["cohort_name"] == "travailleurs_45_plus_manutention_auvergne"
    assert result["summary"]["feasibility_status"] == "feasible"
    assert result["summary"]["source_population_count"] == 4
    assert result["summary"]["included_count"] == 1
    assert [step["output_count"] for step in result["steps"]] == [4, 3, 3, 1]
    assert report_path.exists()


def test_cohort_engine_evaluates_longitudinal_yaml_definition(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    output = tmp_path / "mcdst_tables"
    report_path = tmp_path / "cohort_report_v02.json"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )
    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_review_decisions(proposed["review_queue"]))
    apply_review_workdir(workdir, decisions_path)
    apply_mapping_file(workdir / "mapping_valide.yaml", exports, output)

    result = evaluate_cohort_definition(output, LONGITUDINAL_COHORT_FIXTURE, report_path)

    assert result["cohort_name"] == "manutention_avant_restriction_2024"
    assert result["schema_version"] == "mcdst-cohort-v0.2"
    assert result["summary"]["feasibility_status"] == "feasible"
    assert result["summary"]["included_count"] == 1
    assert result["summary"]["longitudinal_sequences_count"] == 1
    assert result["summary"]["required_tables"] == [
        "conclusion_medicale",
        "exposition_professionnelle",
        "travailleur",
        "visite_sante_travail",
    ]
    assert [step["output_count"] for step in result["steps"]] == [4, 3, 1]
    assert result["steps"][-1]["id"] == "longitudinal:exposition_before_restriction"
    assert result["steps"][-1]["matched_pairs_count"] == 1
    assert report_path.exists()


def test_learning_dataset_export_from_reviewed_mapping(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    dataset_path = tmp_path / "training" / "mapping_dataset.jsonl"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )
    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_review_decisions(proposed["review_queue"]))
    apply_review_workdir(workdir, decisions_path)

    summary = build_learning_dataset(workdir, dataset_path)
    records = [
        json.loads(line)
        for line in dataset_path.read_text(encoding="utf-8").splitlines()
    ]

    assert summary["records_count"] == 105
    assert summary["by_task"] == {"column_mapping": 69, "value_mapping": 36}
    assert summary["by_label"] == {"blocked_s4": 6, "positive": 99}
    assert any(record["label"] == "blocked_s4" and record["source_column"] == "NomUsuel" for record in records)
    assert any(
        record["task"] == "value_mapping"
        and record["source_value"] == "Charge mentale"
        and record["target_value"] == "CHARGE_MENTALE"
        for record in records
    )


def test_source_graph_contains_explicit_join_rules(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )
    mapping = proposed["mapping"]
    join_rules = mapping["join_rules"]

    assert (workdir / "source_graph.json").exists()
    assert (workdir / "join_rules.json").exists()
    assert "&id" not in (workdir / "mapping_propose.yaml").read_text(encoding="utf-8")
    assert "*id" not in (workdir / "mapping_propose.yaml").read_text(encoding="utf-8")
    assert len(proposed["source_graph"]["nodes"]) == 9
    assert len(join_rules) == 31
    assert len([rule for rule in join_rules if rule["status"] == "auto_validable"]) == 12
    assert len([rule for rule in join_rules if rule["status"] == "a_revoir"]) == 19

    worker_rule = find_join_rule(
        join_rules,
        primary_file="export_01_individus.csv",
        foreign_file="export_03_actes.csv",
        column="ClePers",
    )
    assert worker_rule["key_role"] == "travailleur_id"
    assert worker_rule["join_type"] == "primary_foreign_key"
    assert worker_rule["cardinality"] == "one_to_one"

    company_rule = find_join_rule(
        join_rules,
        primary_file="export_02_structures.csv",
        foreign_file="export_01_individus.csv",
        column="CleAdh",
    )
    assert company_rule["key_role"] == "entreprise_id"
    assert company_rule["cardinality"] == "one_to_many"

    peer_rule = [
        rule for rule in join_rules
        if rule["left"]["source_file"] == "export_03_actes.csv"
        and rule["right"]["source_file"] == "export_04_risques.csv"
    ][0]
    assert peer_rule["join_type"] == "peer_key_overlap"
    assert peer_rule["primary"] is None
    assert peer_rule["foreign"] is None


def test_mapping_yaml_contract_for_draft_and_reviewed_artifacts(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )
    assert_mapping_contract(proposed["mapping"], expected_status="draft")

    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_column_decisions(proposed["review_queue"]))
    validated = apply_review_workdir(workdir, decisions_path)

    assert_mapping_contract(validated, expected_status="reviewed")
    assert validated["review_queue"] == []
    assert validated["mapping_version"] == "0.1.0-draft-reviewed"
    assert validated["reviewed_at"]
    assert validated["review_decisions"]["column_mapping_decisions"]
    assert (
        validated["entities"]["conclusion_medicale"]["fields"]["restriction_flag"]["review_status"]
        == "validated_by_human_review"
    )
    assert (
        validated["entities"]["conclusion_medicale"]["fields"]["amenagement_flag"]["review_status"]
        == "validated_by_human_review"
    )


def test_mapping_registry_learns_from_human_review(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    registry_path = tmp_path / "mapping_registry.yaml"
    first_workdir = tmp_path / "first_work"
    second_workdir = tmp_path / "second_work"

    proposed = propose_mapping_workdir(
        exports,
        first_workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
        registry_path=registry_path,
    )
    assert len(proposed["review_queue"]["pending_column_mappings"]) == 2

    decisions_path = first_workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_column_decisions(proposed["review_queue"]))
    validated = apply_review_workdir(first_workdir, decisions_path, registry_path=registry_path)
    assert validated["learning_registry"]["column_mappings_count"] == 2

    registry = load_registry(registry_path)
    assert len(registry["column_mappings"]) == 2
    assert {
        (entry["source_column"], entry["entity"], entry["target_field"])
        for entry in registry["column_mappings"]
    } == {
        ("Adaptation", "conclusion_medicale", "amenagement_flag"),
        ("Reserve", "conclusion_medicale", "restriction_flag"),
    }

    learned = propose_mapping_workdir(
        exports,
        second_workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
        registry_path=registry_path,
    )
    assert learned["review_queue"]["pending_column_mappings"] == []
    assert len(learned["mapping"]["review_queue"]) == 0
    fields = learned["mapping"]["entities"]["conclusion_medicale"]["fields"]
    assert fields["restriction_flag"]["review_status"] == "validated_by_registry"
    assert fields["amenagement_flag"]["review_status"] == "validated_by_registry"
    assert fields["restriction_flag"]["confidence_score"] == 0.995
    assert fields["amenagement_flag"]["confidence_score"] == 0.995
    assert len(learned["review_queue"]["pending_value_mappings"]) == 4


def test_mapping_engine_reads_windows_encoded_csv(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    rewrite_structures_as_cp1252(exports / "export_02_structures.csv")
    workdir = tmp_path / "work"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_CP1252_EXPORT",
        schema_version="mcdst-v0.1",
    )

    structures = [
        profile for profile in proposed["profiles"]
        if profile["file"] == "export_02_structures.csv"
    ][0]
    region = [column for column in structures["columns"] if column["name"] == "Territoire"][0]

    assert structures["row_count"] == 3
    assert "Auvergne-Rhône-Alpes" in region["examples"]


def test_mapping_engine_detects_semicolon_csv_source(tmp_path):
    exports = tmp_path / "exports"
    exports.mkdir()
    copyfile(
        SEMICOLON_CSV_FIXTURE,
        exports / "salaries_suivis_small.csv",
    )

    profiles = profile_exports(exports)
    assert len(profiles) == 1

    profile = profiles[0]
    column_names = [column["name"] for column in profile["columns"]]

    assert profile["row_count"] == 1
    assert len(column_names) > 20
    assert "Matricule personne" in column_names
    assert "Type suivi actuel" in column_names
    assert "Poste actuel" in column_names
    assert "travailleur" in profile["inferred_entities"]


def test_mapping_engine_reads_excel_workbook_sources(tmp_path):
    exports = tmp_path / "exports"
    exports.mkdir()
    write_excel_fixture(exports / "logiciel_spsti.xlsx")
    workdir = tmp_path / "work"
    output = tmp_path / "mcdst_tables"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_EXCEL_EXPORT",
        schema_version="mcdst-v0.1",
    )
    mapping = proposed["mapping"]

    assert len(mapping["entities"]) == 13
    assert len(mapping["blocked_fields"]) == 6
    assert len(mapping["join_candidates"]) == 31
    assert len(mapping["join_rules"]) == 31
    assert all(source["format"] == "excel" for source in mapping["source_files"])
    assert {source["sheet"] for source in mapping["source_files"]} == {
        "export_01_individus",
        "export_02_structures",
        "export_03_actes",
        "export_04_risques",
        "export_05_biometrie",
        "export_06_pathologies_atmp",
        "export_07_arrets",
        "export_08_vaccinations",
        "export_09_duerp",
    }

    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_column_decisions(proposed["review_queue"]))
    apply_review_workdir(workdir, decisions_path)

    state = apply_mapping_file(workdir / "mapping_valide.yaml", exports, output)
    assert state["quality"]["summary"]["generated_tables"]["travailleur"] == 4
    assert state["quality"]["summary"]["generated_tables"]["exposition_professionnelle"] == 4
    assert state["quality"]["summary"]["generated_tables"]["examen_complementaire"] == 4
    assert state["quality"]["summary"]["generated_tables"]["risque_unite_travail"] == 4


def test_mapping_cli_roundtrip(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    output = tmp_path / "mcdst_tables"

    assert main(
        [
            "mapping",
            "propose",
            str(exports),
            "--out",
            str(workdir),
            "--source-system",
            "POC_SPSTI_MULTI_EXPORT",
        ]
    ) == 0

    decisions_path = workdir / "review_decisions.yaml"
    copyfile(FIXTURE_REVIEW_DECISIONS, decisions_path)

    assert main(["mapping", "review", str(decisions_path), "--workdir", str(workdir)]) == 0
    assert main(
        [
            "mapping",
            "apply",
            str(workdir / "mapping_valide.yaml"),
            "--exports",
            str(exports),
            "--out",
            str(output),
        ]
    ) == 0

    assert (workdir / "mapping_propose.yaml").exists()
    assert (workdir / "mapping_valide.yaml").exists()
    assert (output / "travailleur.csv").exists()
    assert (output / "quality_report.json").exists()


def test_cohort_cli_evaluate(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    output = tmp_path / "mcdst_tables"
    report_path = tmp_path / "cohort_report.json"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )
    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_review_decisions(proposed["review_queue"]))
    apply_review_workdir(workdir, decisions_path)
    apply_mapping_file(workdir / "mapping_valide.yaml", exports, output)

    assert main(["cohort", "evaluate", str(COHORT_FIXTURE), "--tables", str(output), "--out", str(report_path)]) == 0
    assert report_path.exists()


def test_learning_cli_dataset(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    dataset_path = tmp_path / "training" / "mapping_dataset.jsonl"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )
    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_review_decisions(proposed["review_queue"]))
    apply_review_workdir(workdir, decisions_path)

    assert main(["learning", "dataset", "--workdir", str(workdir), "--out", str(dataset_path)]) == 0
    assert dataset_path.exists()


def test_learning_cli_train_evaluate_and_predict(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    dataset_path = tmp_path / "training" / "mapping_dataset.jsonl"
    model_path = tmp_path / "training" / "column_model.json"
    suggestions_path = tmp_path / "training" / "mapping_suggestions.json"

    proposed = propose_mapping_workdir(
        exports,
        workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
    )
    decisions_path = workdir / "review_decisions.yaml"
    write_yaml(decisions_path, approve_all_review_decisions(proposed["review_queue"]))
    apply_review_workdir(workdir, decisions_path)
    build_learning_dataset(workdir, dataset_path)

    assert main(["learning", "train", "--dataset", str(dataset_path), "--out", str(model_path)]) == 0
    assert model_path.exists()

    model = json.loads(model_path.read_text(encoding="utf-8"))
    assert model["model_kind"] == "mcdst-column-centroid-v0.1"
    assert model["examples_count"] == 69
    assert model["label_counts"]["__BLOCKED_S4__"] == 6

    assert main(["learning", "evaluate", "--dataset", str(dataset_path), "--model", str(model_path)]) == 0
    suggested_workdir = tmp_path / "suggested_work"
    suggested = propose_mapping_workdir(
        exports,
        suggested_workdir,
        source_system="POC_SPSTI_MULTI_EXPORT",
        schema_version="mcdst-v0.1",
        learning_model_path=model_path,
    )
    assert suggested["learning_suggestions"]["suggestions_count"] >= 69
    assert (suggested_workdir / "mapping_suggestions.json").exists()

    suggestions = build_column_suggestions(workdir, model_path, suggestions_path, top_k=2)
    assert suggestions["blocked_s4_columns"] == 6
    assert suggestions["suggestions_count"] >= 69

    suggestion_payload = json.loads(suggestions_path.read_text(encoding="utf-8"))
    assert any(
        item["source_column"] == "TypeExamen"
        and item["target"] == "examen_complementaire.examen_type_concept_id"
        for item in suggestion_payload["suggestions"]
    )

    assert main(
        [
            "learning",
            "suggest",
            "--workdir",
            str(workdir),
            "--model",
            str(model_path),
            "--out",
            str(suggestions_path),
            "--top-k",
            "2",
        ]
    ) == 0
    assert main(
        [
            "learning",
            "predict",
            "--model",
            str(model_path),
            "--source-file",
            "export_05_biometrie.csv",
            "--source-column",
            "TypeExamen",
            "--source-type",
            "categorical",
            "--source-sensitivity",
            "S3",
            "--entity",
            "examen_complementaire",
            "--example",
            "IMC",
            "--top-k",
            "3",
        ]
    ) == 0


def test_mapping_api_roundtrip(tmp_path):
    exports = copy_fixture_exports(tmp_path / "exports")
    workdir = tmp_path / "work"
    output = tmp_path / "mcdst_tables"

    with running_api() as base_url:
        health = api_get(f"{base_url}/health")
        assert health["status"] == "ok"
        assert "MCD-ST" in api_get_text(f"{base_url}/")
        assert "runPropose" in api_get_text(f"{base_url}/web/app.js")
        registry_path = tmp_path / "api_registry.yaml"

        proposed = api_post(
            f"{base_url}/api/mapping/propose",
            {
                "exports": str(exports),
                "workdir": str(workdir),
                "source_system": "POC_SPSTI_MULTI_EXPORT",
                "schema_version": "mcdst-v0.1",
                "registry_path": str(registry_path),
            },
        )
        assert proposed["status"] == "proposed"
        assert proposed["summary"]["entities"] == 13
        assert proposed["summary"]["join_rules"] == 31
        assert proposed["summary"]["review_columns"] == 2
        assert proposed["summary"]["review_values"] == 4
        assert proposed["summary"]["registry_column_mappings"] == 0
        assert proposed["summary"]["learning_suggestions"] == 0
        assert (workdir / "mapping_propose.yaml").exists()

        artifact = api_get(f"{base_url}/api/artifact?path={workdir / 'mapping_propose.yaml'}")
        assert "mapping_id:" in artifact["content"]
        saved = api_post(
            f"{base_url}/api/artifact",
            {
                "path": str(workdir / "ui_note.yaml"),
                "content": "status: checked\n",
            },
        )
        assert saved["status"] == "saved"
        assert (workdir / "ui_note.yaml").read_text(encoding="utf-8") == "status: checked\n"

        review_queue = api_get(f"{base_url}/api/mapping/review-queue?workdir={workdir}")
        reviewed = api_post(
            f"{base_url}/api/mapping/review",
            {
                "workdir": str(workdir),
                "registry_path": str(registry_path),
                "decisions": approve_all_review_decisions(review_queue),
            },
        )
        assert reviewed["status"] == "reviewed"
        assert reviewed["summary"]["review_columns"] == 0
        assert reviewed["summary"]["review_values"] == 0
        assert reviewed["summary"]["registry_column_mappings"] == 2
        assert (workdir / "mapping_valide.yaml").exists()
        assert registry_path.exists()

        learned_workdir = tmp_path / "learned_work"
        learned = api_post(
            f"{base_url}/api/mapping/propose",
            {
                "exports": str(exports),
                "workdir": str(learned_workdir),
                "source_system": "POC_SPSTI_MULTI_EXPORT",
                "schema_version": "mcdst-v0.1",
                "registry_path": str(registry_path),
            },
        )
        assert learned["summary"]["review_columns"] == 0
        assert learned["summary"]["registry_column_mappings"] == 2

        applied = api_post(
            f"{base_url}/api/mapping/apply",
            {
                "mapping": str(workdir / "mapping_valide.yaml"),
                "exports": str(exports),
                "out": str(output),
            },
        )
        assert applied["status"] == "applied"
        assert applied["summary"]["generated_tables"]["travailleur"] == 4
        assert applied["summary"]["generated_tables"]["examen_complementaire"] == 4
        assert (output / "quality_report.json").exists()

        cohort_report = tmp_path / "cohort_report.json"
        cohort = api_post(
            f"{base_url}/api/cohort/evaluate",
            {
                "tables": str(output),
                "definition": str(COHORT_FIXTURE),
                "out": str(cohort_report),
            },
        )
        assert cohort["status"] == "evaluated"
        assert cohort["summary"]["included_count"] == 1
        assert cohort["steps"][-1]["id"] == "criteria:exposure"
        assert cohort_report.exists()

        missing = api_get_error(f"{base_url}/api/mapping/review-queue")
        assert missing["status"] == "error"
        assert "workdir" in missing["message"]


def copy_fixture_exports(path: Path) -> Path:
    path.mkdir()
    for filename in CANONICAL_EXPORT_FILES:
        copyfile(FIXTURE_EXPORTS / filename, path / filename)
    return path


def rewrite_structures_as_cp1252(path: Path) -> None:
    path.write_bytes(
        (
            "CleAdh,Site,Adherent,Siret,APE,Nb,Territoire\n"
            "A100,E10,Clinique,12345678900011,8610Z,180,Auvergne-Rhône-Alpes\n"
            "A101,E11,Bois SARL,98765432100022,4332A,22,Auvergne-Rhône-Alpes\n"
            "A102,E12,Logistique Sud,11122233300044,5229B,8,Auvergne-Rhône-Alpes\n"
        ).encode("cp1252")
    )


def write_excel_fixture(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for filename in CANONICAL_EXPORT_FILES:
        csv_path = FIXTURE_EXPORTS / filename
        sheet = workbook.create_sheet(csv_path.stem)
        with csv_path.open(newline="", encoding="utf-8") as file:
            for row_index, row in enumerate(csv.reader(file), start=1):
                for column_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=column_index, value=value)
    workbook.save(path)


def approve_all_column_decisions(review_queue: dict) -> dict:
    return {
        "column_mapping_decisions": [
            {
                "id": item["id"],
                "action": "approve",
                "source_file": item["source_file"],
                "source_column": item["source_column"],
                "entity": item["entity"],
                "target_field": item["target_field"],
                "transform": item["transform"],
                "reviewer": "test",
                "reason": "Approved in acceptance test.",
            }
            for item in review_queue["pending_column_mappings"]
        ],
        "value_mapping_decisions": [],
    }


def approve_all_review_decisions(review_queue: dict) -> dict:
    decisions = approve_all_column_decisions(review_queue)
    decisions["value_mapping_decisions"] = [
        {
            "id": item["id"],
            "action": "approve",
            "source_file": group["source_file"],
            "source_column": group["source_column"],
            "entity": group["entity"],
            "target_field": group["target_field"],
            "source_value": item["source_value"],
            "target_value": item["target_value"],
            "reviewer": "test",
            "reason": "Approved in acceptance test.",
        }
        for group in review_queue["pending_value_mappings"]
        for item in group["mappings"]
        if item["status"] == "a_revoir"
    ]
    return decisions


def find_join_rule(join_rules: list[dict], *, primary_file: str, foreign_file: str, column: str) -> dict:
    matches = [
        rule for rule in join_rules
        if rule["primary"]
        and rule["foreign"]
        and rule["primary"]["source_file"] == primary_file
        and rule["foreign"]["source_file"] == foreign_file
        and rule["primary"]["column"] == column
        and rule["foreign"]["column"] == column
    ]
    assert len(matches) == 1
    return matches[0]


def assert_mapping_contract(mapping: dict, *, expected_status: str) -> None:
    assert set(mapping) >= MAPPING_CONTRACT_KEYS
    assert mapping["schema_version"] == "mcdst-v0.1"
    assert mapping["review_status"] == expected_status
    assert set(mapping["thresholds"]) == {"auto_validable", "a_revoir"}
    assert mapping["source_files"]
    assert mapping["join_candidates"]
    assert mapping["join_rules"]
    assert mapping["entities"]
    assert mapping["blocked_fields"]

    for entity in mapping["entities"].values():
        assert set(entity) >= ENTITY_CONTRACT_KEYS
        assert ".csv" in entity["source_file"] or ".xlsx#" in entity["source_file"]
        assert entity["review_status"] in {"auto_draft", "reviewed"}
        assert entity["fields"]
        for field in entity["fields"].values():
            assert set(field) >= FIELD_CONTRACT_KEYS
            assert 0 <= field["confidence_score"] <= 1
            assert field["sensitivity"] in {"S1", "S2", "S3"}
            assert field["review_status"] in {
                "auto_validable",
                "a_revoir",
                "validated_by_registry",
                "validated_by_human_review",
            }

    for blocked in mapping["blocked_fields"]:
        assert blocked["sensitivity"] == "S4"
        assert blocked["recommended_action"] == "exclude_from_mcd"


@contextmanager
def running_api():
    server = create_server("127.0.0.1", 0)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        yield f"http://127.0.0.1:{server.server_port}"
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)


def api_get(url: str) -> dict:
    with urlopen(url, timeout=5) as response:
        return json.loads(response.read().decode("utf-8"))


def api_get_text(url: str) -> str:
    with urlopen(url, timeout=5) as response:
        return response.read().decode("utf-8")


def api_post(url: str, payload: dict) -> dict:
    request = Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urlopen(request, timeout=10) as response:
        return json.loads(response.read().decode("utf-8"))


def api_get_error(url: str) -> dict:
    try:
        return api_get(url)
    except HTTPError as exc:
        return json.loads(exc.read().decode("utf-8"))
