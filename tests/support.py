from __future__ import annotations

import csv
import json
import threading
from contextlib import contextmanager
from pathlib import Path
from shutil import copyfile
from urllib.error import HTTPError
from urllib.request import Request, urlopen

from mcdst.api import create_server


FIXTURE_EXPORTS = Path(__file__).parent / "fixtures" / "mapping_poc_exports"
SEMICOLON_CSV_FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "semicolon_csv_export"
    / "salaries_suivis_small.csv"
)
COHORT_FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "cohorts"
    / "travailleurs_45_plus_manutention.yaml"
)
LONGITUDINAL_COHORT_FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "cohorts"
    / "manutention_avant_restriction.yaml"
)
ARRET_REPRISE_COHORT_FIXTURE = (
    Path(__file__).parent
    / "fixtures"
    / "cohorts"
    / "arret_avant_visite_reprise.yaml"
)
FIXTURE_REVIEW_DECISIONS = Path(__file__).parent / "fixtures" / "mapping_poc_review_decisions.yaml"
CANONICAL_EXPORT_FILES = (
    "export_01_individus.csv",
    "export_02_structures.csv",
    "export_03_actes.csv",
    "export_04_risques.csv",
    "export_05_biometrie.csv",
    "export_06_pathologies_atmp.csv",
    "export_07_arrets.csv",
    "export_08_vaccinations.csv",
    "export_09_duerp.csv",
)
S4_SOURCE_COLUMNS = {"NomUsuel", "Prenom", "TelPortable"}
SYNTHETIC_DIRECT_IDENTIFIER_VALUES = {
    "Martin",
    "Aline",
    "Bernard",
    "Karim",
    "Dubois",
    "Nadia",
    "Leroy",
    "Marc",
    "0600000001",
    "0600000002",
    "0600000003",
    "0600000004",
}
MAPPING_CONTRACT_KEYS = {
    "mapping_id",
    "source_system",
    "schema_version",
    "mapping_version",
    "review_status",
    "generated_at",
    "thresholds",
    "source_files",
    "join_candidates",
    "join_rules",
    "alternate_entity_sources",
    "entities",
    "review_queue",
    "value_mappings",
    "blocked_fields",
}
ENTITY_CONTRACT_KEYS = {"source_file", "review_status", "fields"}
FIELD_CONTRACT_KEYS = {
    "source",
    "confidence_score",
    "sensitivity",
    "transform",
    "review_status",
}


def copy_fixture_exports(path: Path) -> Path:
    path.mkdir()
    for filename in CANONICAL_EXPORT_FILES:
        copyfile(FIXTURE_EXPORTS / filename, path / filename)
    return path


def rewrite_structures_as_cp1252(path: Path) -> None:
    path.write_bytes(
        (
            "CleAdh,Site,Adherent,Siret,APE,Nb,Territoire\n"
            "A100,E10,Clinique,12345678900011,8610Z,180,Auvergne-Rhône-Alpes\n"
            "A101,E11,Bois SARL,98765432100022,4332A,22,Auvergne-Rhône-Alpes\n"
            "A102,E12,Logistique Sud,11122233300044,5229B,8,Auvergne-Rhône-Alpes\n"
        ).encode("cp1252")
    )


def write_excel_fixture(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for filename in CANONICAL_EXPORT_FILES:
        csv_path = FIXTURE_EXPORTS / filename
        sheet = workbook.create_sheet(csv_path.stem)
        with csv_path.open(newline="", encoding="utf-8") as file:
            for row_index, row in enumerate(csv.reader(file), start=1):
                for column_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=column_index, value=value)
    workbook.save(path)


def approve_all_column_decisions(review_queue: dict) -> dict:
    return {
        "column_mapping_decisions": [
            {
                "id": item["id"],
                "action": "approve",
                "source_file": item["source_file"],
                "source_column": item["source_column"],
                "entity": item["entity"],
                "target_field": item["target_field"],
                "transform": item["transform"],
                "reviewer": "test",
                "reason": "Approved in acceptance test.",
            }
            for item in review_queue["pending_column_mappings"]
        ],
        "value_mapping_decisions": [],
        "join_rule_decisions": [],
    }


def approve_no_review_decisions() -> dict:
    return {
        "column_mapping_decisions": [],
        "value_mapping_decisions": [],
        "join_rule_decisions": [],
    }


def approve_all_review_decisions(review_queue: dict) -> dict:
    decisions = approve_all_column_decisions(review_queue)
    decisions["value_mapping_decisions"] = [
        {
            "id": item["id"],
            "action": "approve",
            "source_file": group["source_file"],
            "source_column": group["source_column"],
            "entity": group["entity"],
            "target_field": group["target_field"],
            "source_value": item["source_value"],
            "target_value": item["target_value"],
            "reviewer": "test",
            "reason": "Approved in acceptance test.",
        }
        for group in review_queue["pending_value_mappings"]
        for item in group["mappings"]
        if item["status"] == "a_revoir"
    ]
    decisions["join_rule_decisions"] = [
        {
            "id": item["id"],
            "action": "approve",
            "key_role": item["key_role"],
            "join_type": item["join_type"],
            "reviewer": "test",
            "reason": "Approved in acceptance test.",
        }
        for item in review_queue.get("pending_join_rules", [])
    ]
    return decisions


def find_join_rule(join_rules: list[dict], *, primary_file: str, foreign_file: str, column: str) -> dict:
    matches = [
        rule for rule in join_rules
        if rule["primary"]
        and rule["foreign"]
        and rule["primary"]["source_file"] == primary_file
        and rule["foreign"]["source_file"] == foreign_file
        and rule["primary"]["column"] == column
        and rule["foreign"]["column"] == column
    ]
    assert len(matches) == 1
    return matches[0]


def assert_mapping_contract(mapping: dict, *, expected_status: str) -> None:
    assert set(mapping) >= MAPPING_CONTRACT_KEYS
    assert mapping["schema_version"] == "mcdst-v0.1"
    assert mapping["review_status"] == expected_status
    assert set(mapping["thresholds"]) == {"auto_validable", "a_revoir"}
    assert mapping["source_files"]
    assert mapping["join_candidates"]
    assert mapping["join_rules"]
    assert mapping["entities"]
    assert mapping["blocked_fields"]

    for entity in mapping["entities"].values():
        assert set(entity) >= ENTITY_CONTRACT_KEYS
        assert ".csv" in entity["source_file"] or ".xlsx#" in entity["source_file"]
        assert entity["review_status"] in {"auto_draft", "reviewed"}
        assert entity["fields"]
        for field in entity["fields"].values():
            assert set(field) >= FIELD_CONTRACT_KEYS
            assert 0 <= field["confidence_score"] <= 1
            assert field["sensitivity"] in {"S1", "S2", "S3"}
            assert field["review_status"] in {
                "auto_validable",
                "a_revoir",
                "validated_by_registry",
                "validated_by_human_review",
            }

    for blocked in mapping["blocked_fields"]:
        assert blocked["sensitivity"] == "S4"
        assert blocked["recommended_action"] == "exclude_from_mcd"


@contextmanager
def running_api():
    server = create_server("127.0.0.1", 0)
    thread = threading.Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        yield f"http://127.0.0.1:{server.server_port}"
    finally:
        server.shutdown()
        server.server_close()
        thread.join(timeout=2)


def api_get(url: str) -> dict:
    with urlopen(url, timeout=5) as response:
        return json.loads(response.read().decode("utf-8"))


def api_get_text(url: str) -> str:
    with urlopen(url, timeout=5) as response:
        return response.read().decode("utf-8")


def api_post(url: str, payload: dict) -> dict:
    request = Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urlopen(request, timeout=10) as response:
        return json.loads(response.read().decode("utf-8"))


def api_get_error(url: str) -> dict:
    try:
        return api_get(url)
    except HTTPError as exc:
        return json.loads(exc.read().decode("utf-8"))
